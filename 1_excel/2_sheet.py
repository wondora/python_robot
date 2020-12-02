from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff"

ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet", 2) #2번째 index에 sheet 생성, index는 0부터

new_ws = wb["NewSheet"] #Dict 형태로 sheet에 접근

print(wb.sheetnames) #모든 sheet 이름 확인

#sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")