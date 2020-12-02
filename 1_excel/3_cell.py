from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])
print(ws["A1"].value)

for x in range(1,11):
    for y in range(1,11):
        ws.cell(row=x, column=y, value=randint(0,100))

wb.save("sample.xlsx")
