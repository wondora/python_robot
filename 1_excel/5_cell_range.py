from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

ws.append(["번호", "영어", "수학"])
for i in range(1,11):
    ws.append([i, randint(0,100), randint(0,100)])

# col_B = ws["B"]

# # for cell in col_B:
# #     print(cell.value)

# col_range = ws["B:C"]

# # for cols in col_range:
# #     for cell in cols:
# #         print(cell.value)

# row_title = ws[1] #첫번째 row만 가져오기
# row_range2 = ws[2:6] #2~6번줄까지 가져오기 slice하고는 다르다. 5까지가 아니다.!!!!

# #셀정보 가져오기
# from openpyxl.utils.cell import coordinate_from_string
# for rows in row_range2:
#     for cell in rows:
#         # print(cell.coordinate, end=" ")
#         xy = coordinate_from_string(cell.coordinate)
#         print(xy, end=" ")
#         #xy[0] 는 A, xy[1]는 2 이다....!!        
#     print()


# row_range2 = ws[2:ws.max_row] #2번째 줄부터 마지막 줄까지

# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row)

# for row in ws.iter_rows(): #위와 똑같은 결과 나온다. 범위로 지정가능
#     print(row[2].value)

# for row in ws.iter_rows(min_row=1, max_row=4):
#     print(row[2].value)

# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
#     print(row[0].value, row[1].value)


for col in ws.iter_cols(min_row=2): # max_row, min_col, max_col 다 적을 필요없다.
    print(col)




# print(tuple(ws.columns))

wb.save("sample.xlsx")